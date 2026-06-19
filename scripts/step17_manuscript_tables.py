"""
Step 17: Manuscript Tables (SOP Phase 12 supplement)
Generates publication-ready Tables 1–4 as CSV + styled Excel.
"""

import time, warnings
import numpy as np
import pandas as pd
from pathlib import Path

warnings.filterwarnings("ignore")

BASE_DIR  = Path("/home/preonath/Desktop/Preonath_Project/Zika_Dengue")
PROC_DIR  = BASE_DIR / "01_processed_data" / "deg_tables"
RES_DIR   = BASE_DIR / "03_results"
OUT_DIR   = BASE_DIR / "05_manuscript_tables"
LOG_FILE  = BASE_DIR / "logs" / "step17_tables.log"

for d in [OUT_DIR, LOG_FILE.parent]:
    d.mkdir(parents=True, exist_ok=True)

def log(msg):
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, "a") as f:
        f.write(line + "\n")


def make_table1():
    """Table 1: Dataset summary (all datasets used in this study)."""
    log("Generating Table 1: Dataset summary ...")
    rows = [
        {
            "GEO Accession":    "GSE110496",
            "Reference":        "Zanini et al. 2018",
            "Cell/Tissue Type": "Huh7 (hepatoma cells)",
            "Virus":            "DENV-2 (16681), ZIKV (PRVABC59)",
            "Platform":         "10x Genomics scRNA-seq",
            "Timepoints":       "4h, 12h, 24h, 48h",
            "MOI":              "DENV: MOI 1 & 3; ZIKV: MOI 1",
            "N samples":        "12 pseudobulk (4 timepoints × 3 conditions)",
            "Analysis role":    "Discovery (primary)",
            "DEGs":             "DENV=527; ZIKV=176",
        },
        {
            "GEO Accession":    "GSE118305",
            "Reference":        "Schmid et al. 2018",
            "Cell/Tissue Type": "HMDM (human monocyte-derived macrophages)",
            "Virus":            "ZIKV (FSSADE, Puerto Rico)",
            "Platform":         "RNA-seq (HOMER FPKM)",
            "Timepoints":       "12h, 18h, 24h",
            "MOI":              "Not specified",
            "N samples":        "5 ZIKV-infected, 5 mock (24h)",
            "Analysis role":    "Validation (macrophage, ZIKV)",
            "DEGs":             "580 (348 up, 232 down at 24h)",
        },
        {
            "GEO Accession":    "GSE94892",
            "Reference":        "Cheung et al. 2017",
            "Cell/Tissue Type": "PBMCs (human peripheral blood)",
            "Virus":            "DENV (patient PBMCs)",
            "Platform":         "Cufflinks RNA-seq",
            "Timepoints":       "Cross-sectional",
            "MOI":              "Patient sample (not applicable)",
            "N samples":        "DENV vs Control",
            "Analysis role":    "Validation (DENV, blood cells)",
            "DEGs":             "116 (91 up, 25 down)",
        },
        {
            "GEO Accession":    "GSE78711",
            "Reference":        "Onorati et al. 2016",
            "Cell/Tissue Type": "Neural progenitor cells (hNPCs)",
            "Virus":            "ZIKV (MR766)",
            "Platform":         "RNA-seq (pre-computed FC)",
            "Timepoints":       "72h post-infection",
            "MOI":              "MOI 0.1",
            "N samples":        "ZIKV vs Mock",
            "Analysis role":    "Validation (ZIKV, neural context)",
            "DEGs":             "1,197 (403 up, 794 down)",
        },
    ]
    t1 = pd.DataFrame(rows)
    t1.to_csv(OUT_DIR / "Table1_Dataset_Summary.csv", index=False)
    log(f"  Saved Table1_Dataset_Summary.csv ({len(t1)} rows)")
    return t1


def make_table2():
    """Table 2: 15 Shared Upregulated DEGs — all annotation layers."""
    log("Generating Table 2: Shared DEG full annotation ...")

    shared = pd.read_csv(RES_DIR / "phase3_shared_degs" / "shared_DEGs_annotated.csv")
    shared_up = shared[shared["log2FC_DENV"] > 0].copy()

    # miRNA targeting info
    mirna_hits = pd.read_csv(RES_DIR / "phase6_mirna" / "mirna_55set_hits_miRTarBase.csv")
    mirna_genes_55 = set()
    for _, row in mirna_hits.iterrows():
        for g in str(row["Genes"]).split(";"):
            mirna_genes_55.add(g.strip())

    # CREBRF miRNA count
    crebrf_mirna_count = len(mirna_hits[mirna_hits["Genes"].str.contains("CREBRF", na=False)])

    # Validation info
    val_genes   = {"CREBRF", "INHBE", "RND1", "TSPYL2"}
    mirna_hubs  = {"CREBRF", "SIRT4", "TSPYL2"}

    # Pathway info from KEGG results
    kegg = pd.read_csv(RES_DIR / "phase4_pathways" / "enrichr_shared_up_KEGG.csv")
    gene_pathways = {}
    for _, row in kegg[kegg["Adj. P-value"] < 0.1].iterrows():
        for g in str(row.get("Genes", "")).split(";"):
            g = g.strip()
            if g:
                gene_pathways.setdefault(g, []).append(row["Term"].split(" (")[0][:25])

    # Build full table
    rows = []
    for _, r in shared_up.sort_values("log2FC_DENV", ascending=False).iterrows():
        g = r["symbol"]
        rows.append({
            "Gene Symbol":         g,
            "Gene Name":           r.get("name", ""),
            "log2FC (DENV)":       round(r["log2FC_DENV"], 3),
            "padj (DENV)":         f"{r['padj_DENV']:.2e}",
            "log2FC (ZIKV)":       round(r["log2FC_ZIKV"], 3),
            "padj (ZIKV)":         f"{r['padj_ZIKV']:.2e}",
            "Host factor class":   "Novel" if not (r.get("is_proviral") or r.get("is_antiviral")) else
                                   ("Proviral" if r.get("is_proviral") else "Antiviral"),
            "ISG":                 "Yes" if r.get("is_ISG") else "No",
            "55-set miRNA target": "Yes" if g in mirna_genes_55 else "No",
            "miRNA targeting count": crebrf_mirna_count if g == "CREBRF" else
                                      len(mirna_hits[mirna_hits["Genes"].str.contains(g, na=False)]),
            "NPC replicated":      "Yes (p=1.18e-4)" if g in val_genes else "No",
            "Macrophage replicated": "No",
            "DENV PBMC replicated": "No",
            "Convergent pathways": "; ".join(gene_pathways.get(g, ["—"])),
            "Priority candidate":  "★★★" if (g in val_genes and g in mirna_hubs) else
                                   ("★★" if g in val_genes else ("★" if g in mirna_hubs else "")),
        })

    t2 = pd.DataFrame(rows)
    t2.to_csv(OUT_DIR / "Table2_Shared_DEGs_Annotated.csv", index=False)
    log(f"  Saved Table2_Shared_DEGs_Annotated.csv ({len(t2)} rows × {len(t2.columns)} cols)")
    return t2


def make_table3():
    """Table 3: Multi-layer intersection summary."""
    log("Generating Table 3: Multi-layer intersection ...")

    shared_up = pd.read_csv(RES_DIR / "phase3_shared_degs" / "shared_DEGs_annotated.csv")
    shared_up = shared_up[shared_up["log2FC_DENV"] > 0]

    val_genes_npc  = {"CREBRF", "INHBE", "RND1", "TSPYL2"}
    mirna_targets  = {"CREBRF", "SIRT4", "TSPYL2", "CREBRF"}

    rows = []
    for g in shared_up["symbol"].tolist():
        layer_a = True  # all are in shared DEG set
        layer_b = g in val_genes_npc       # NPC validation
        layer_c = g in mirna_targets       # miRNA targeting
        layer_d = g in {"CXCL1", "BIRC3"} # STRING network hub
        n_layers = sum([layer_a, layer_b, layer_c, layer_d])
        rows.append({
            "Gene":           g,
            "Layer 1 (Shared DEG)":     "✓",
            "Layer 2 (NPC Replicated)": "✓" if layer_b else "",
            "Layer 3 (miRNA Target)":   "✓" if layer_c else "",
            "Layer 4 (Network Hub)":    "✓" if layer_d else "",
            "Total Layers":  n_layers,
            "Classification": "Top-tier" if n_layers >= 3 else
                               ("Strong" if n_layers == 2 else "Candidate"),
        })
    t3 = pd.DataFrame(rows).sort_values(["Total Layers", "Gene"], ascending=[False, True])
    t3.to_csv(OUT_DIR / "Table3_MultiLayer_Intersection.csv", index=False)
    log(f"  Saved Table3_MultiLayer_Intersection.csv")
    return t3


def make_table4():
    """Table 4: Convergent pathways (enriched in shared DEGs)."""
    log("Generating Table 4: Convergent pathways ...")

    kegg   = pd.read_csv(RES_DIR / "phase4_pathways" / "enrichr_shared_up_KEGG.csv")
    gobp   = pd.read_csv(RES_DIR / "phase4_pathways" / "enrichr_shared_up_GO_BP.csv")
    hall   = pd.read_csv(RES_DIR / "phase4_pathways" / "enrichr_shared_up_Hallmarks.csv")
    react  = pd.read_csv(RES_DIR / "phase4_pathways" / "enrichr_shared_up_Reactome.csv")

    def sig_rows(df, db):
        df2 = df[df["Adj. P-value"] < 0.1].copy()
        df2["Database"] = db
        df2["Gene Count"] = df2["Overlap"].str.split("/").str[0].astype(int)
        return df2[["Database", "Term", "Gene Count", "Odds Ratio", "P.value", "Adj. P-value", "Genes"]]

    all_pathways = pd.concat([
        sig_rows(kegg,  "KEGG"),
        sig_rows(gobp,  "GO:BP"),
        sig_rows(hall,  "MSigDB Hallmarks"),
        sig_rows(react, "Reactome"),
    ], ignore_index=True)

    all_pathways = all_pathways.sort_values("Adj. P-value")
    all_pathways["Odds Ratio"] = all_pathways["Odds Ratio"].round(2)
    all_pathways["P.value"]    = all_pathways["P.value"].map(lambda x: f"{x:.2e}")
    all_pathways["Adj. P-value"] = all_pathways["Adj. P-value"].map(lambda x: f"{x:.3f}")
    all_pathways["Term"] = all_pathways["Term"].str[:60]

    all_pathways.to_csv(OUT_DIR / "Table4_Convergent_Pathways.csv", index=False)
    log(f"  Saved Table4_Convergent_Pathways.csv ({len(all_pathways)} significant pathways)")
    return all_pathways


def make_excel_workbook(t1, t2, t3, t4):
    """Combine all 4 tables into a single styled Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        def write_sheet(wb, name, df, header_color="1F3864"):
            ws = wb.create_sheet(name)
            header_fill = PatternFill("solid", fgColor=header_color)
            header_font = Font(color="FFFFFF", bold=True)
            thin = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
            for j, col in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=j, value=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, horizontal="center")
                cell.border = thin

            for i, row_data in enumerate(df.values, 2):
                for j, val in enumerate(row_data, 1):
                    cell = ws.cell(row=i, column=j, value=val)
                    cell.border = thin
                    cell.alignment = Alignment(wrap_text=False, horizontal="center")
                    # Zebra rows
                    if i % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor="F5F5F5")

            # Auto column width
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)
            return ws

        write_sheet(wb, "Table 1 - Datasets",     t1, "1A237E")
        write_sheet(wb, "Table 2 - Shared DEGs",  t2, "1B5E20")
        write_sheet(wb, "Table 3 - MultiLayer",   t3, "4A148C")
        write_sheet(wb, "Table 4 - Pathways",     t4, "BF360C")

        out_path = OUT_DIR / "Supplementary_Tables.xlsx"
        wb.save(out_path)
        log(f"  Saved Supplementary_Tables.xlsx (4 sheets)")
    except ImportError:
        log("  openpyxl not available — skipping Excel; CSV tables still saved")


def main():
    log("=" * 60)
    log("Step 17: Manuscript Tables (Phase 12)")
    log("=" * 60)

    t1 = make_table1()
    t2 = make_table2()
    t3 = make_table3()
    t4 = make_table4()
    make_excel_workbook(t1, t2, t3, t4)

    log("\n" + "=" * 60)
    log("STEP 17 COMPLETE — Manuscript Tables")
    log("Output directory: " + str(OUT_DIR))
    log("Tables generated:")
    log("  Table1_Dataset_Summary.csv          (4 datasets)")
    log("  Table2_Shared_DEGs_Annotated.csv    (15 genes × all annotation layers)")
    log("  Table3_MultiLayer_Intersection.csv  (4-layer evidence scoring)")
    log("  Table4_Convergent_Pathways.csv      (significant enriched pathways)")
    log("  Supplementary_Tables.xlsx           (all 4 tables, styled)")


if __name__ == "__main__":
    main()
